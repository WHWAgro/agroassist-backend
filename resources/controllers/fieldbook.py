from flask import Response, request,send_file
import pandas as pd
from flask_restful import Resource
import json
from datetime import datetime, timedelta
#import pandas as pd
#from get_project_root import root_path
from resources.errors import  InternalServerError
from database.models.Program import ProgramClass,db
from resources.services.programServices import *
from flask_jwt_extended import jwt_required,get_jwt_identity
import xlsxwriter
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from copy import copy





class FieldBookApi(Resource):
  

  @jwt_required()
  def get(self):
    try:
        response={}
        response['status']=200
        response['message']=0
        user_id =  get_jwt_identity()
      

        data={}
        
        
        user_company=getUserCompanies(user_id)
  
        
        

        
        
        companies="( "
        companiesNumeric=[]
        for i,company in enumerate(user_company):
            if i==0:
                companies=companies+str(company["_id"])+","
                companiesNumeric.append(str(company["_id"]))
            else:
                if company['visible']:
                    companies=companies+str(company["_id"])+","
                    companiesNumeric.append(str(company["_id"]))
            
        companies = companies[:-1]
        companies=companies+" )"
        markets = request.args.get('markets').split(",")
        species=request.args.get('species')

        

        #species_format="( "
        #for specie in species:
            
         #   species_format=species_format+str(specie)+","
        
        #species_format = species_format[:-1]
        #species_format=species_format+" )"

        markets_format="( "
        for market in markets:
            
            markets_format=markets_format+str(market)+","
        markets_format = markets_format[:-1]
        markets_format=markets_format+" )"
        

    
        
        programs=getProgramsMarketFilter(user_id,companies,markets_format)

        progrmas_format="( "
        for program in programs:
            progrmas_format=progrmas_format+str(program["_id"])+","
        progrmas_format = progrmas_format[:-1]
        progrmas_format=progrmas_format+" )"
        
        
        print("cuadernos de campo")
        print(progrmas_format)
        print(companies)
        print(species)
        fields = getFieldMarketFilter(progrmas_format,companies,species)

        legacy_fieldbooks = {
    "2": [
        "ALTO LOS OLMOS",
        "ALTO NILAHUE",
        "ASQUE",
        "COIQUE",
        "COLCHAMAULE",
        "CUN CUN",
        "DAÑICALQUI",
        "DEMAIHUE",
        "EL CASTILLO",
        "EL PATAGUAL",
        "EL TORREON",
        "EUSKADI",
        "HIGUERA ORIENTE",
        "HUIFQUENCO",
        "LA AGUADA",
        "LA TORRE",
        "LAS VIÑAS",
        "LOS PINOS",
        "LUIS URETA",
        "PALQUIBUDIS",
        "PAREDONES ABAJO",
        "PUMANQUE",
        "ROBERTO FARIAS",
        "SAN SEBASTIAN",
        "SANTA RITA",
        "SUIZA",
        "VICENTE JEREZ",
    ],
    "6": [
        "ANDES BERRIES",
        "COIQUE",
        "COLCHAMAULE",
        "EL AMANECER",
        "EL PALENQUE",
        "FRAMPARQUE",
        "IDELIO BECERRA",
        "JOSE LUIS YAÑEZ LEON",
        "LAS VERTIENTES",
        "LOS PINOS",
        "SANTA ISABEL",
    ],
    "4": [
        "AGRICOLA OCTAVIO BUSTOS",
        "AGROFARIAS EL CONVENTO",
        "CIELO ABIERTO",
        "COLCHAMAULE",
        "EL CARMEN",
        "EL COPIHUE",
        "EL ERMITAÑO",
        "EL MAITEN/ SAN SEBASTIAN",
        "LAS VIÑAS",
        "LOS PINOS",
        "LUIS URETA",
        "PALQUIBUDIS",
        "PAREDONES ABAJO",
        "PUMANQUE",
        "ROBERTO FARIAS",
        "SAN SEBASTIAN",
        "SANTA RITA",
        "SUIZA",
        "VICENTE JEREZ",
    ],
    "5": [
        "COLCHAMAULE",
        "EL CARMEN",
        "LA UNION",
        "PARCELA 10-13",
        "PARCELA 22",
    ],
        }
        
        
        if (("1" in companiesNumeric or "25" in companiesNumeric or "30" in companiesNumeric) and (str(species) in legacy_fieldbooks) ):
            print ("entro aca")
            fieldsToAdd = legacy_fieldbooks[str(species)]

            for i,field in enumerate(fieldsToAdd):
               
                fields.append({"_id":-i,"sag_code":str(field),"field_name":""})
            
            
        
        if fields== False:
            response['status']=400 
            response['message']=1
        

        data={}
        data['fields']=fields
        response['data']=data
        
        if response.get('status') == 200:

            return {'response': response}, 200
        
        else: 
            
            return {'response': response}, 400

    except Exception as e:
        print(e)
        response['message']=2
        return {'response': response},500
  
  
  
  
  

  
  
  
class FieldBookFullApi(Resource):
  

  @jwt_required()
  def get(self):
    try:
        response={}
        response['status']=200
        response['message']=0
        user_id =  get_jwt_identity()
      

        data={}
        
        
        user_company=getUserCompanies(user_id)

        companies="( "
        for company in user_company:
            companies=companies+str(company["_id"])+","
       
        #esta seccion consigue el valor de markets
        markets_dict=getTableDict("market")
        markets_final=""
        companies = companies[:-1]
        companies=companies+" )"
        markets = request.args.get('markets').split(",")
       
        markets_format="( "
             
        for market in markets:
            if market=='':
                markets_final='Todos  '
                continue
            markets_final=markets_final+str(markets_dict[int(market)]["market_name"])+", "
            markets_format=markets_format+str(market)+","
       
        
        markets_format = markets_format[:-1]
        markets_format=markets_format+" )"
        markets_final = markets_final[:-2]


       
      
        
        #---------------------
        

       
        print('----&&&&&&&&&&&&&7comeinzo&&&&&&&&&&&&&&&&&&&&&&&----')
       
        
        ##programs=getProgramsMarketFilter(user_id,companies,markets_format)

        #progrmas_format="( "
        #for program in programs:
        #    progrmas_format=progrmas_format+str(program["_id"])+","
        #progrmas_format = progrmas_format[:-1]
        #progrmas_format=progrmas_format+" )"
        
        
        #print("programs")
        #print(programs)
        #fields_aux = getFieldMarketFilter(progrmas_format)
        
        ##----especies
        
        dict_species=getTableDict("species")
        species= request.args.get('species').split(",")
        species_id= request.args.get('species')

        species_final=""
        species_format="( "
        for specie in species:
            species_final=species_final+str(dict_species[int(specie)]["species_name"])+", "
            species_format=species_format+str(specie)+","
       
        species_format = species_format[:-1]
        species_format=species_format+" )"
        species_final = species_final[:-2]
        #----------------------

       
        fields = request.args.get('fields').split(",")

        print(request.args)
        if fields== False:
            response['status']=400 
            response['message']=1
        

        data={}
        data['fields']=fields
        response['data']=data


       
        
        
        fields_format="( "
        for field in fields:
            fields_format=fields_format+str(field)+","
        
        fields_format = fields_format[:-1]
        
        fields_format=fields_format+" )"
       

        field_book_data_full=getFieldBookDataFull(fields_format,species_format)

    
        
        

        field_weather_location=[]
        
        
        field_fb={
        }
        for field in fields:
            field_fb[str(field)]={"data":[],"company":"","field_name":"","CSG_code":"","location":"","size":0,"weather_location":""}
            

        products=getTableDict("Products")
        objectives=getTableDict("Objectives")
        machinery=getTableDict("machinery")
        workers=getTableDict("workers")
        users=getTableDict("users")
        
        
        
        checked_fields=[]
        checked_plots=[]
        for row in field_book_data_full:
            if row['f_id'] not in checked_fields:
                checked_fields.append(row['f_id'])
                field_fb[str(row["f_id"])]["company"]=row["company_name"]
                field_fb[str(row["f_id"])]["field_name"]=row["field_name"]
                field_fb[str(row["f_id"])]["CSG_code"]=row["sag_code"]
                field_fb[str(row["f_id"])]["location"]=row["locat"]
                field_fb[str(row["f_id"])]["weather_location"]=row["f_w_location"]
                field_weather_location.append(row["f_w_location"])
            if row['plot_id'] not in checked_plots:
                print(row)
                checked_plots.append(row['plot_id'])
                field_fb[str(row["f_id"])]["size"]=row["plot_size"]+field_fb[str(row["f_id"])]["size"]
        print("hola")
        print(field_weather_location)
        print("----------*********")
        f_w_location_format="( "
        for field in field_weather_location:
             f_w_location_format= f_w_location_format+str(field)+","
        
        f_w_location_format=  f_w_location_format[:-1]
        
        f_w_location_format= f_w_location_format+" )"

        print("hola")
        print(f_w_location_format)
        print("----------********1")

        field_weather=[]
        
        if f_w_location_format != "( )":
            field_weather=getFieldWeather(f_w_location_format)
        print("----------********2")

        for row in field_book_data_full:
            

            if row["plot_task_status_id"] !=2 or row["t_o_objectives"] ==None:
                continue
            print(row)

            row["t_o_objectives"]= ast.literal_eval(row["t_o_objectives"])
            row["t_o_products"]= ast.literal_eval(row["t_o_products"])
            row["t_o_ingredients"]= ast.literal_eval(row["t_o_ingredients"])
            row["t_o_dosage"]= ast.literal_eval(row["t_o_dosage"])
            row["t_o_dosage_unit"]= ast.literal_eval(row["t_o_dosage_unit"])
            row["t_o_tractor"]= ast.literal_eval(row["t_o_tractor"])
            row["t_o_sprayer"]= ast.literal_eval(row["t_o_sprayer"])
            row["t_o_operators"]= ast.literal_eval(row["t_o_operators"])
            row["t_o_dosage_responsible"]= ast.literal_eval(row["t_o_dosage_responsible"])

            
            for index,elem in enumerate(row["t_o_objectives"]):
                print(elem)
                wetting=int(row["t_o_wetting"])
                objetivo=elem
                producto=row["t_o_products"][index]
                ingrediente=row["t_o_ingredients"][index]
                
                
                if can_cast_to_number(elem):
                    objetivo=objectives[elem]["objective_name"]
                    
                if can_cast_to_number(producto):
                    producto=products[producto]["product_name"]
                    
                if can_cast_to_number(ingrediente):
                    ingrediente=products[ingrediente]["chemical_compounds"]
                    
                print("productos")

                dosage_unit=row["t_o_dosage_unit"][index]
                dosage=row["t_o_dosage"][index]

                unit_dosage=""
                

                unit_hectare=""
                dosage_hectare=0
            
            
                dosage=float(dosage)
                print("or6")

                if dosage_unit == 1:
                    print(" 1")
                    unit=" gr"
                    unit_dosage="gr/100L"
                   
                    unit_hectare="gr/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 9:
                    print(" 1")
                    unit=" gr Ingrediente Activo"
                    unit_dosage="gr Ingrediente Activo/100L"
                    
                    unit_hectare="gr Ingrediente Activo /Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 2:
                    print(" 2")
                    unit=" Kg"
                    unit_dosage="Kg/100L"
                    
                    unit_hectare="Kg/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 10:
                    print(" 2")
                    unit=" Kg Ingrediente Activo"
                    unit_dosage="Kg Ingrediente Activo/100L"
                    
                    unit_hectare="Kg Ingrediente Activo/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 3:
                    print(" 3")
                    unit=" gr"
                    unit_dosage="gr/100L"
                    
                    
                    unit_hectare="gr/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                elif dosage_unit == 4:
                    print(" 4")
                    unit=" Kg"
                    unit_dosage="Kg/100L"
                    
                    
                    unit_hectare="Kg/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                if dosage_unit == 5:
                    print(" 5")
                    unit=" cc"
                    unit_dosage="cc/100L"
                   
                   
                    unit_hectare="cc/Há"
                    dosage_hectare=dosage*(wetting/100)
                if dosage_unit == 11:
                    print(" 5")
                    unit=" cc Ingrediente Activo"
                    unit_dosage="cc Ingrediente Activo/100L"
                   
                   
                    unit_hectare="cc Ingrediente Activo/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 6:
                    print(" 6")
                    unit=" L"
                    unit_dosage="L/100L"
                   
                    unit_hectare="L/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 12:
                    print(" 6")
                    unit=" L Ingrediente Activo"
                    unit_dosage="L Ingrediente Activo/100L"
                    unit_hectare="L Ingrediente Activo/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 7:
                    print(" 7")
                    unit=" cc"
                    unit_dosage="cc/100L"
                    unit_hectare="cc/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                elif dosage_unit == 8:
                    print(" 8")
                    unit=" L"
                    unit_dosage="L/100L"
                    unit_hectare="L/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                print("cuartel 1")
                dosage='{:,.2f}'.format(float(dosage)).replace(',','*').replace('.', ',').replace('*','.').replace(',00','')+" "+unit_dosage 
                print("cuartel 2")
                dosage_hectare='{:,.2f}'.format(float(dosage_hectare)).replace(',','*').replace('.', ',').replace('*','.').replace(',00','')+" "+unit_hectare
                    

                application_date=row["t_o_application_date"]

                print(application_date)
                print(row["t_o_phi"])

                harvest_date=application_date+timedelta(days=int(row["t_o_phi"])+1)

                method="Bomba de espalda"
                maquina=""
                tractor=""
                responsable=users[row['program_id_user']]['user_name']
                
                print('hola')
                print('')
                if row["t_o_application_method"]  in [1,3]:
                    for m in row["t_o_sprayer"]:
                        maquina=maquina+machinery[m]["name"]+";"
                    maquina=maquina[:-1]
                    for m in row["t_o_tractor"]:
                        tractor=''
                        if m != None and m !='None':

                            tractor=tractor+machinery[m]["name"]+";"
                        else:
                            tractor='No Usado'
                    tractor=tractor[:-1]
                    method="Nebulizador"
                    if row["t_o_application_method"] == 3:
                        method="Otro"
                else:
                    for m in row["t_o_tractor"]:
                        maquina=maquina+machinery[m]["name"]+";"
                    maquina=maquina[:-1]
                print("hola")
                aplicador=""
                dosificador=""
                print(row["t_o_dosage_responsible"])
                print('fdf')
                for d in row["t_o_dosage_responsible"]:
                    print(workers[d])
                    dosificador=dosificador+workers[d]["name"]+";"
                dosificador=dosificador[:-1]

                print('chao')

                for d in row["t_o_operators"]:
                    print(workers[d])
                    aplicador=aplicador+workers[d]["name"]+";"
                aplicador=aplicador[:-1]
                print('fsdfdf')
                

                

                weather={"temperature_min":"","temperature_max":"","wind":"","wind_direction":"","humidity":""}
                for weather_day in field_weather:
                    if weather_day["date"]==application_date.strftime("%Y-%m-%d") and weather_day["weather_locations_id"]==row["f_w_location"]:
                        weather=weather_day
                        break
            
                field_fb[str(row["f_id"])]["data"].append({
                                                                "Fecha Aplicación":application_date.strftime("%Y-%m-%d"),
                                                                "Hora Inicio":row["t_o_time_start"],
                                                                "Hora Fin":row["t_o_time_end"],
                                                                "Cuartel":row["plot_name"],
                                                                "Superficie (Hás)":str(row["plot_size"]),
                                                                "Especie":str(dict_species[int(row["plot_species_id"])]["species_name"]),
                                                                "Variedad":row["plot_variety"],
                                                                "Objetivo":objetivo,
                                                                "Producto Comercial":producto,
                                                                "Ingrediente Activo":ingrediente,
                                                                "Dosis/100 L":dosage,
                                                                "Dosis/Há":dosage_hectare,
                                                                "Mojamiento (L/Há )":wetting,
                                                                "Volumen Total (L)":str(int(row["t_o_volumen"])),
                                                                "Horas Reingreso":str(row["t_o_reentry"]),
                                                                "Dias Carencia":str(row["t_o_phi"]),
                                                                "Fecha Viable de Cosecha":harvest_date.strftime("%Y-%m-%d"),
                                                                "Método de Aplicación":method,
                                                                "Codigo Máquina":maquina,
                                                                "Tractor":tractor,
                                                                "Nombre Dosificador":dosificador,
                                                                "Nombre Aplicador":aplicador,
                                                                'Responsable técnico':responsable,
                                                                "T mínima (C°)":weather["temperature_min"],
                                                                "T máxima (C°)":weather["temperature_max"],
                                                                "Humedad (%)":weather["humidity"],
                                                                "Vel Viento (km/h)":weather["wind"],
                                                                "Dir Viento":weather["wind_direction"]

                                                                
                                                                })
        
        print("----------------")
        
        
        
        
                
            
            


        
        # Write dataframes to Excel
         # Write the dataframe to an Excel file
        myuuid = uuid.uuid4()
        doc_name = "files/fieldbooks/"+str(myuuid)+".xlsx"

        template_path = "files/fieldbooks/tmp_exp.xlsx"
        wb = Workbook()
    
        # Access the active worksheet (assuming there's only one sheet)
        original_sheet = wb.active
        images = []
        start_row = 6  # For example, starting from row 4
        for img in original_sheet._images:
            images.append(img)

        # Create a new workbook
        legacy_wb = load_workbook("files/fieldbooks/nutrisco_historico.xlsx")
        
        legacy_fieldbooks = {
    "2": [
        "ALTO LOS OLMOS",
        "ALTO NILAHUE",
        "ASQUE",
        "COIQUE",
        "COLCHAMAULE",
        "CUN CUN",
        "DAÑICALQUI",
        "DEMAIHUE",
        "EL CASTILLO",
        "EL PATAGUAL",
        "EL TORREON",
        "EUSKADI",
        "HIGUERA ORIENTE",
        "HUIFQUENCO",
        "LA AGUADA",
        "LA TORRE",
        "LAS VIÑAS",
        "LOS PINOS",
        "LUIS URETA",
        "PALQUIBUDIS",
        "PAREDONES ABAJO",
        "PUMANQUE",
        "ROBERTO FARIAS",
        "SAN SEBASTIAN",
        "SANTA RITA",
        "SUIZA",
        "VICENTE JEREZ",
    ],
    "6": [
        "ANDES BERRIES",
        "COIQUE",
        "COLCHAMAULE",
        "EL AMANECER",
        "EL PALENQUE",
        "FRAMPARQUE",
        "IDELIO BECERRA",
        "JOSE LUIS YAÑEZ LEON",
        "LAS VERTIENTES",
        "LOS PINOS",
        "SANTA ISABEL",
    ],
    "4": [
        "AGRICOLA OCTAVIO BUSTOS",
        "AGROFARIAS EL CONVENTO",
        "CIELO ABIERTO",
        "COLCHAMAULE",
        "EL CARMEN",
        "EL COPIHUE",
        "EL ERMITAÑO",
        "EL MAITEN/ SAN SEBASTIAN",
        "LAS VIÑAS",
        "LOS PINOS",
        "LUIS URETA",
        "PALQUIBUDIS",
        "PAREDONES ABAJO",
        "PUMANQUE",
        "ROBERTO FARIAS",
        "SAN SEBASTIAN",
        "SANTA RITA",
        "SUIZA",
        "VICENTE JEREZ",
    ],
    "5": [
        "COLCHAMAULE",
        "EL CARMEN",
        "LA UNION",
        "PARCELA 10-13",
        "PARCELA 22",
    ]}
        
        for idx, field_id in enumerate(list(field_fb.keys()), start=0):
            print("new_sheet")
            
            print(field_id)
            thick_border = Border(left=Side(style='medium'), 
                      right=Side(style='medium'), 
                      top=Side(style='medium'), 
                      bottom=Side(style='medium'))
            
            if int(field_id) <=0:
                
                positive_field_id = int(field_id) * -1
                if species_id in legacy_fieldbooks:
                    src_ws = legacy_wb[legacy_fieldbooks[species_id][positive_field_id]]
                    dest_ws = wb.create_sheet(legacy_fieldbooks[species_id][positive_field_id])
                    print("Loaded sheets:", wb.sheetnames)

                    for row in src_ws.iter_rows():
                        for cell in row:
                            new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                orig_border = copy(cell.border)

                                new_border = Border(
                                    left=Side(style=("medium" if orig_border.left and orig_border.left.style == "thin" else orig_border.left.style),
                                            color=orig_border.left.color if orig_border.left else None),
                                    right=Side(style=("medium" if orig_border.right and orig_border.right.style == "thin" else orig_border.right.style),
                                            color=orig_border.right.color if orig_border.right else None),
                                    top=Side(style=("medium" if orig_border.top and orig_border.top.style == "thin" else orig_border.top.style),
                                            color=orig_border.top.color if orig_border.top else None),
                                    bottom=Side(style=("medium" if orig_border.bottom and orig_border.bottom.style == "thin" else orig_border.bottom.style),
                                                color=orig_border.bottom.color if orig_border.bottom else None),
                                )

                                new_cell.border = new_border
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                


                    autofit_columns(dest_ws)  
                print("skip")
                continue
            field_data=field_fb[field_id]
            if len(field_data["data"])==0:
                print('este campo no tiene ODAs')

                field_data["data"].append({"Fecha Aplicación":"No hay ODAs Completadas",
                                           
                                                                
                                                                "Hora Inicio":"",
                                                                "Hora Fin":"",
                                                                "Cuartel":"",
                                                                "Superficie (Hás)":"",
                                                                "Especie":"",
                                                                "Variedad":"",
                                                                "Objetivo":"",
                                                                "Producto Comercial":"",
                                                                "Ingrediente Activo":"",
                                                                "Dosis/100 L":"",
                                                                "Dosis/Há":"",
                                                                "Mojamiento (L/Há )":"",
                                                                "Volumen Total (L)":"",
                                                                "Horas Reingreso":"",
                                                                "Dias Carencia":"",
                                                                "Fecha Viable de Cosecha":"",
                                                                "Método de Aplicación":"",
                                                                "Codigo Máquina":"",
                                                                "Tractor":"",
                                                                "Nombre Dosificador":"",
                                                                "Nombre Aplicador":"",
                                                                'Responsable técnico':"",
                                                                "T mínima (C°)":"",
                                                                "T máxima (C°)":"",
                                                                "Humedad (%)":"",
                                                                "Vel Viento (km/h)":"",
                                                                "Dir Viento":""
                                                            })

            df=pd.DataFrame(field_data["data"])
            new_sheet = wb.copy_worksheet(original_sheet)
            new_sheet.title = f'Datos {field_data["CSG_code"]}' 
            
            

            from datetime import datetime

            # Get the current date
            current_date = datetime.now()

            # Format the date as "day-month-year"
            date_string = current_date.strftime("%d-%m-%Y")

            
            
            
            

            print("------ret1")
            

            
            

            print("------ret2")

            cell=new_sheet.cell(row=start_row+0, column=1, value="Razón Social: ")
            cell.border=thick_border
            
            
            cell=new_sheet.cell(row=start_row+0, column=2, value=field_data["company"])
            cell.border = thick_border
            print("------ret3")

            cell=new_sheet.cell(row=start_row+1, column=1, value="Nombre Predio: ")
            cell.border=thick_border
            
            
            cell=new_sheet.cell(row=start_row+1, column=2, value=field_data["field_name"])
            cell.border = thick_border

            
            
            
            
            
            print("------ret4")
            cell=new_sheet.cell(row=start_row+2, column=1, value="Código SAG predio (CSG): ")
            cell.border=thick_border
            
            
            cell=new_sheet.cell(row=start_row+2, column=2, value=field_data["CSG_code"])
            cell.border = thick_border
            
            print("------ret5")
            cell=new_sheet.cell(row=start_row+3, column=1, value="Especies: ")
            cell.border=thick_border
            
            cell=new_sheet.cell(row=start_row+3, column=2, value=species_final)
            cell.border = thick_border
            
           
            print("------ret6")
            cell=new_sheet.cell(row=start_row+4, column=1, value="Región:")
            cell.border=thick_border
            
            cell=new_sheet.cell(row=start_row+4, column=2, value="VI")
            cell.border = thick_border
            
            
            print("------ret7")
            cell=new_sheet.cell(row=start_row+5, column=1, value="Comuna: ")
            cell.border=thick_border
            
            cell=new_sheet.cell(row=start_row+5, column=2, value=field_data["location"])
            cell.border = thick_border
            print("------ret8")
            cell=new_sheet.cell(row=start_row+6, column=1, value="Paises a Exportar: ")
            cell.border=thick_border
            
           
            cell=new_sheet.cell(row=start_row+6, column=2, value=markets_final)
            cell.border = thick_border
            print("------ret9")
           
            cell=new_sheet.cell(row=start_row+7, column=1, value="Superficie Total: ")
            cell.border=thick_border
            
            
            cell=new_sheet.cell(row=start_row+7, column=2, value=str(field_data["size"])+' Há')
            cell.border = thick_border
            
            print("----datos inciales")
            

            
            
            
            
            light_border = Border( 
                      
                       
                      bottom=Side(style='thin'))
            rows = dataframe_to_rows(df)
            row_n=start_row+9
            r_idx=row_n
            for  row in rows:  #starts at 3 as you want to skip the first 2 rows
                print("Hola")
                print(row)
                
                
                if row==[None]:
                    print("chao")
                    continue
                if len(row)>2 and row[1]=='No hay ODAs Completadas':
                    r_idx=r_idx+1
                r_idx=r_idx+1
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        continue
                    
                    cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)
                    # Apply outer border to the cell
                    if r_idx == start_row + 10 :
                        cell.border = thick_border
            autofit_columns(new_sheet) 

            
      
            
            
        print("file closing")
        wb.remove(original_sheet)
        print("file closing")

            # Write title and subtitles
    

        wb.save(doc_name)

           

        
        

        # Send the Excel file as a response
        return send_file(doc_name, as_attachment=True)
        
        if response.get('status') == 200:

            return {'response': response}, 200
        
        else: 
            
            return {'response': response}, 400

    except Exception as e:
        print(e)
        response['message']=2
        return {'response': response},500
  
  
  
  
class FieldBookExportApi(Resource):
  

  @jwt_required()
  def get(self):
    try:
        response={}
        response['status']=200
        response['message']=0
        user_id =  get_jwt_identity()
      

        data={}
        
        
        user_company=getUserCompanies(user_id)
        
        companies="( "
        for company in user_company:
            companies=companies+str(company["_id"])+","
        companies = companies[:-1]
        companies=companies+" )"
        markets = request.args.get('markets').split(",")
        print("hola0")
        markets_format="( "
        for market in markets:
            markets_format=markets_format+str(market)+","
        markets_format = markets_format[:-1]
        markets_format=markets_format+" )"
        print(companies)
        print(markets_format)
        print("hola1")
        
        ##programs=getProgramsMarketFilter(user_id,companies,markets_format)

        #progrmas_format="( "
        #for program in programs:
        #    progrmas_format=progrmas_format+str(program["_id"])+","
        #progrmas_format = progrmas_format[:-1]
        #progrmas_format=progrmas_format+" )"
        
        
        #print("programs")
        #print(programs)
        #fields_aux = getFieldMarketFilter(progrmas_format)
        fields = request.args.get('fields').split(",")
        
        
        if fields== False:
            response['status']=400 
            response['message']=1
        

        data={}
        data['fields']=fields
        response['data']=data


        data1 = {'tasks': ['test1', 'test2', 'test3'],
            }
        df1 = pd.DataFrame(data1)
        data2 = {'tasks': ['test5', 'test3', 'test4'],
            }
        df2 = pd.DataFrame(data2)

        dfs=[df1,df2]

        print("fields format")
        fields_format="( "
        for field in fields:
            fields_format=fields_format+str(field)+","
        
        fields_format = fields_format[:-1]
        
        fields_format=fields_format+" )"
       

        field_book_data=getFieldBookData(fields_format)
        
        
        field_fb={
        }
        for field in fields:
            field_fb[str(field)]={"data":[],"company":"","CSG_code":"","location":"","varieties":[]}

        products=getTableDict("Products")
        objectives=getTableDict("Objectives")
        processed_rows=[]
        p_r=[]
        for row in field_book_data:
            row_id=str(row["_id"])+'-'+str(row["f_id"])+'-'+str(row["to_id"])
            processed=row_id  in processed_rows
            
            
            if processed ==False:
               
             
                
            
                processed_rows.append(row_id)
                field_fb[str(row["f_id"])]["company"]=row["company_name"]
                field_fb[str(row["f_id"])]["CSG_code"]=row["sag_code"]
                field_fb[str(row["f_id"])]["location"]=row["locat"]
            if row["application_date"] is None or row["id_status"]!=2 or row_id in p_r  :
                if row["id_status"]==2:
                    field_fb[str(row["f_id"])]["varieties"].append(row["variety"])
                continue
            if row["id_status"]==2:
                field_fb[str(row["f_id"])]["varieties"].append(row["variety"])
                p_r.append(row_id)

        ##----
           
            products_id=ast.literal_eval(row["id_product"])
            dppus=ast.literal_eval(row["dosage_parts_per_unit"])
            dosages=ast.literal_eval(row["dosage"])
            date_start=row["date_start"].strftime("%d-%m-%Y")
            
            application_date=row["application_date"].strftime("%d-%m-%Y")
            
            wetting=row["wetting"]

            for i in range(0,len(products_id)):
                product_id=products_id[i][0]
                product=products[product_id]
                product_name=product["product_name"]
                active_ingredient=product["chemical_compounds"]
                dosage=float(dosages[i][0])
                dosage_unit=dppus[i][0]
                unit_dosage=""
                
                
                
                if dosage_unit == 1:
                    print(" 1")
                    unit=" gr"
                    unit_dosage="gr/100L"
                    
                    unit_hectare="gr/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 2:
                    
                    unit=" Kg"
                    unit_dosage="Kg/100L"
                    
                    unit_hectare="Kg/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 3:
                    print(" 3")
                    unit=" gr"
                    unit_dosage="gr/100L"
                    
                    unit_hectare="gr/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                elif dosage_unit == 4:
                    print(" 4")
                    unit=" Kg"
                    unit_dosage="Kg/100L"
                    
                    
                    unit_hectare="Kg/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                if dosage_unit == 5:
                    print(" 5")
                    unit=" cc"
                    unit_dosage="cc/100L"
                    
                    unit_hectare="cc/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 6:
                    print(" 6")
                    unit=" L"
                    unit_dosage="L/100L"
                    
                    unit_hectare="L/Há"
                    dosage_hectare=dosage*(wetting/100)
                elif dosage_unit == 7:
                    print(" 7")
                    unit=" cc"
                    unit_dosage="cc/100L"
                    
                    
                    unit_hectare="cc/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                elif dosage_unit == 8:
                    print(" 8")
                    unit=" L"
                    unit_dosage="L/100L"
                    
                    unit_hectare="L/Há"
                    dosage_hectare=dosage
                    dosage=str(dosage/(wetting/100))
                dosage='{:,.2f}'.format(float(dosage)).replace(',','*').replace('.', ',').replace('*','.')


                objective=objectives[product["id_objective"]]["objective_name"]
                field_fb[str(row["f_id"])]["data"].append({"Objetivo":objective,
                                                            "Nombre Producto":product_name,
                                                            "Fecha":application_date,
                                                            "Ingredientes Activos":active_ingredient,
                                                            "Dosis/100L":dosage+unit})
                
                
            
            


        title = pd.DataFrame({'Title': ['Protocolo de Exportación']})
        subtitle = pd.DataFrame({'Información': ['Campos:'+fields_format, 'Mercados: '+markets_format]})
        
        # Write dataframes to Excel
         # Write the dataframe to an Excel file
        myuuid = uuid.uuid4()
        doc_name = "files/fieldbooks/"+str(myuuid)+".xlsx"

        template_path = "files/fieldbooks/tmp_exp.xlsx"
        wb = load_workbook(template_path)
    
        # Access the active worksheet (assuming there's only one sheet)
        original_sheet = wb.active
        images = []
        start_row = 14  # For example, starting from row 4
        for img in original_sheet._images:
            images.append(img)

        # Create a new workbook
        
        
        for idx, field_id in enumerate(list(field_fb.keys()), start=0):
            print("new_sheet")
            print(idx)
            

            field_data=field_fb[field_id]
            if len(field_data["data"])==0:

                field_data["data"].append({"Objetivo":"No hay ODAs Completadas",
                                                            "Nombre Producto":"",
                                                            "Fecha":"",
                                                            "Ingredientes Activos":"",
                                                            "Dosis/100L":""})
            df=pd.DataFrame(field_data["data"])
            new_sheet = wb.copy_worksheet(original_sheet)
            new_sheet.title = f'Datos {field_data["CSG_code"]}' 
            
            print("hola")

            from datetime import datetime

            # Get the current date
            current_date = datetime.now()

            # Format the date as "day-month-year"
            date_string = current_date.strftime("%d-%m-%Y")

            varieties=set(field_data["varieties"])


            new_sheet.cell(row=start_row+1, column=2, value="Fecha: "+ date_string)
            new_sheet.cell(row=start_row+2, column=2, value="Razón Social: "+field_data["company"])
            new_sheet.cell(row=start_row+3, column=2, value="Variedades: "+', '.join(varieties))
            new_sheet.cell(row=start_row+4, column=2, value="Código CSG: "+field_data["CSG_code"])
            
            thick_border = Border(left=Side(style='medium'), 
                      right=Side(style='medium'), 
                      top=Side(style='medium'), 
                      bottom=Side(style='medium'))
            light_border = Border( 
                      
                       
                      bottom=Side(style='thin'))
            rows = dataframe_to_rows(df)
            row_n=start_row+7
            for r_idx, row in enumerate(rows, start_row+6):  
                row_n=row_n+1
                for c_idx, value in enumerate(row, 1):
                    if c_idx==1:
                        continue
                    
                    cell = new_sheet.cell(row=r_idx, column=c_idx, value=value)
                    # Apply outer border to the cell
                    if r_idx == start_row + 6 or r_idx == row_n or c_idx == 1:
                        cell.border = thick_border
            new_sheet.cell(row=row_n, column=2, value="Certifica")
            for c_idx in range(2, 7):
                cell = new_sheet.cell(row=row_n+4, column=c_idx)
                cell.border = light_border

            new_sheet.cell(row=row_n+5, column=2, value="Nombre Profesiona o Ténico")
            new_sheet.cell(row=row_n+5, column=5, value="Firma")

            new_sheet.cell(row=row_n+7, column=2, value="Nota: - Formato referencial")
            


        legacy_wb = load_workbook("files/fieldbooks/nutrisco_historico.xlsx")
        src_ws = legacy_wb["CUN CUN"]
        dest_ws = wb.create_sheet("Copied_Sheet1")
        print("Loaded sheets:", wb.sheetnames)

        for row in src_ws.iter_rows():
            for cell in row:
                new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font
                    new_cell.border = cell.border
                    new_cell.fill = cell.fill
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection
                    new_cell.alignment = cell.alignment
    
            
            
        print("file closing")
        wb.remove(original_sheet)
        print("file closing")

            # Write title and subtitles
    

        wb.save(doc_name)

           

        
        

        # Send the Excel file as a response
        return send_file(doc_name, as_attachment=True)
        
        if response.get('status') == 200:

            return {'response': response}, 200
        
        else: 
            
            return {'response': response}, 400

    except Exception as e:
        print(e)
        response['message']=2
        return {'response': response},500
  
  
  
  
  

  


  
